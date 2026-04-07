from typing import Any
from dataclasses import dataclass
from openpyxl import load_workbook
from django.db import transaction
from client_app.models import Client, Contract, DeliveryAddress
from repositories import client_repository


@dataclass
class Record:
    contract_id: str
    phone: str
    addresses: list[str]


def upload_data(filename: str, binary_file_data: Any) -> None:
    save_file(filename, binary_file_data)
    extracted_data = extract_data_from_file(filename)
    save_data(extracted_data)


@transaction.atomic
def save_data(record: Record) -> None:
    client = client_repository.fetch_client_by_name(record.phone)
    if client is not None:
        return
    client = Client.objects.create(name=record.phone, is_active=True)
    contract = client_repository.fetch_contract_by_name(name=record.contract_id)
    if contract is not None:
        return
    contract = Contract.objects.create(
        name=record.contract_id, client=client, is_active=True
    )
    for address in record.addresses:
        DeliveryAddress.objects.create(name=address, contract=contract)


def save_file(filename: str, binary_file_data: Any) -> None:
    with open(file=filename, mode="wb+") as destination:
        for chunk in binary_file_data.chunks():
            destination.write(chunk)


def extract_data_from_file(filename: str) -> Record:
    workbook = load_workbook(filename=filename, data_only=True)
    sheet = workbook.active
    contract_id = str(sheet.cell(row=2, column=2).value)
    phone = str(sheet.cell(row=9, column=2).value)
    if phone[1] != "+":
        phone = "+" + phone
    addresses = []
    for row in sheet.iter_rows(values_only=True, min_row=12):
        if row is None:
            continue
        addr = row[1]
        if addr is None:
            continue
        addresses.append(str(addr))
    return Record(contract_id=contract_id, phone=phone, addresses=addresses)
