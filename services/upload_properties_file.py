from typing import Any
from dataclasses import dataclass
from openpyxl import load_workbook
from catalog_app.models import PropertyRecord
from repositories import good_repository, property_repository


@dataclass
class Record:
    name: str
    art: str
    design_id: str
    prop1: str
    prop2: str
    prop3: str
    prop4: str
    prop5: str
    prop6: str
    prop7: str
    prop8: str
    prop9: str
    prop10: str


def upload_data(filename: str, data: Any) -> None:
    save_file(filename, data)
    data = fetch_data_from_file(filename)
    title = data[0]
    to_create = []
    for _, record in enumerate(data):
        if _ == 0:
            continue
        good = good_repository.fetch_good_by_art(record.art)
        if not good:
            continue
        for i in range(1, 11):
            if getattr(record, f"prop{i}"):
                to_create.append(
                    PropertyRecord(
                        good=good,
                        sort_ordering=i,
                        name=getattr(title, f"prop{i}"),
                        value=getattr(record, f"prop{i}"),
                    )
                )
    property_repository.create_properties(to_create)


def save_file(filename: str, data: Any) -> None:
    with open(file=filename, mode="wb+") as destination:
        for chunk in data.chunks():
            destination.write(chunk)


def fetch_data_from_file(filename: str) -> list[Record]:
    data: list[Record] = []
    workbook = load_workbook(filename=filename, data_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True, min_row=2):
        data.append(Record(*row))
    return data
