from typing import Any
from openpyxl import Workbook
from repositories import good_repository
import io


def fetch_goods_to_xlsx_data() -> Any:
    wb = Workbook()
    ws = wb.active
    ws.title = "goods"

    # Заголовок таблицы
    ws.cell(row=1, column=1, value="Код")
    ws.cell(row=1, column=2, value="Артикул")
    ws.cell(row=1, column=3, value="Наименование")
    ws.cell(row=1, column=4, value="Активный")

    # Данные таблицы
    for i, good in enumerate(good_repository.fetch_all_goods(), 2):
        ws.cell(row=i, column=1, value=good.code)
        ws.cell(row=i, column=2, value=good.art)
        ws.cell(row=i, column=3, value=good.name)
        ws.cell(row=i, column=4, value=good.is_active)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
